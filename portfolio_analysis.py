"""
投资组合分析模块
- 四账户偏离度检测
- 海外指数估值（标普500 PE/PB/股息率 + 中证港股指数）
"""

import os
import re
import warnings
import urllib.request
from datetime import datetime

import numpy as np
import openpyxl

warnings.filterwarnings("ignore")

# ============ 常量 ============

EXCEL_PATH = os.path.expanduser("~/code/skyebee/Investment/current.xlsx")
PROXY_URL = os.environ.get("HTTPS_PROXY", "http://127.0.0.1:10810")

# 四账户目标配比
TARGET_ALLOCATIONS = {
    "安全账户": 0.30,
    "稳定现金流": 0.25,
    "成长账户": 0.35,
    "机会账户": 0.10,
}

# multpl.com 数据源（S&P 500）
MULTPL_URLS = {
    "PE": "https://www.multpl.com/s-p-500-pe-ratio/table/m",
    "PB": "https://www.multpl.com/s-p-500-price-to-book/table/m",
    "股息率": "https://www.multpl.com/s-p-500-dividend-yield/table/m",
}

# 中证港股指数代码
CSINDEX_HK_MAP = {
    "中国互联网": "H11136",
    "内地民营": "H11152",
    "沪港深通金融": "H30550",
}

# 账户行范围映射（基于 current.xlsx 结构）
ACCOUNT_ROW_RANGES = {
    "安全账户": (2, 14),
    "稳定现金流": (15, 19),
    "成长账户": (21, 29),
    "机会账户": (30, 31),
}

# 估值等级分档（与 index_valuation.py 一致）
VALUATION_GRADES = [
    (0, 10, "极度低估 🥶"),
    (10, 20, "低估 🟢"),
    (20, 40, "偏低 🟡"),
    (40, 60, "合理 🟠"),
    (60, 80, "偏高 🔴"),
    (80, 90, "高估 🔥"),
    (90, 101, "极度高估 🚨"),
]


def _get_valuation_grade(percentile: float) -> str:
    """根据百分位返回估值等级"""
    for lo, hi, label in VALUATION_GRADES:
        if lo <= percentile < hi:
            return label
    return "未知"


def _calc_percentile(values: list, current: float) -> float:
    """计算 current 在 values 中的历史分位（0-100）"""
    arr = np.array(values)
    return round(float(np.sum(arr <= current) / len(arr) * 100), 1)


def _parse_multpl_date(date_str: str) -> str:
    """将 'Apr 23, 2026' 转为 '2026-04-23'"""
    date_str = date_str.strip().replace("†", "").strip()
    try:
        dt = datetime.strptime(date_str, "%b %d, %Y")
        return dt.strftime("%Y-%m-%d")
    except ValueError:
        return date_str


# ============ 1. 持仓数据读取 ============


def read_portfolio_from_excel(excel_path: str = None) -> dict:
    """
    从 current.xlsx 读取持仓数据。

    Returns:
        {
            "funds": [{"name": str, "account": str, "subcategory": str, "amount": float}, ...],
            "accounts": {"安全账户": {"amount": float, "items": [...]}, ...},
            "total": float
        }
    """
    path = excel_path or EXCEL_PATH
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"]

    funds = []
    current_account = "未知"

    for row_idx in range(2, ws.max_row + 1):
        # A=类别, F=子类1, H=子类2, I=投资项目, J=数目(金额)
        category = ws.cell(row=row_idx, column=1).value  # A
        subcat1 = ws.cell(row=row_idx, column=6).value or ""  # F
        subcat2 = ws.cell(row=row_idx, column=8).value or ""  # H
        project = ws.cell(row=row_idx, column=9).value  # I
        amount = ws.cell(row=row_idx, column=10).value  # J

        # 更新当前账户
        if category:
            cat_str = str(category).strip()
            if "安全" in cat_str:
                current_account = "安全账户"
            elif "稳定" in cat_str:
                current_account = "稳定现金流"
            elif "成长" in cat_str:
                current_account = "成长账户"
            elif "机会" in cat_str:
                current_account = "机会账户"

        # 跳过无金额或汇总行
        if not amount or not str(amount).replace(".", "").replace("-", "").isdigit():
            continue
        amount_val = float(amount)
        if amount_val <= 0:
            continue

        # 确定子类
        subcategory = str(subcat2).strip() if subcat2 else str(subcat1).strip() if subcat1 else ""
        if not subcategory and project:
            subcategory = str(project).strip()

        item_name = str(project).strip() if project else subcategory

        funds.append({
            "name": item_name,
            "account": current_account,
            "subcategory": subcategory,
            "amount": amount_val,
            "row": row_idx,
        })

    wb.close()

    # 按账户汇总
    accounts = {}
    for acc in TARGET_ALLOCATIONS:
        accounts[acc] = {"amount": 0, "items": []}

    for f in funds:
        acc = f["account"]
        if acc in accounts:
            accounts[acc]["amount"] += f["amount"]
            accounts[acc]["items"].append(f)

    total = sum(a["amount"] for a in accounts.values())

    return {"funds": funds, "accounts": accounts, "total": total}


# ============ 2. 偏离度计算 ============


def calculate_deviation(portfolio_data: dict) -> dict:
    """
    计算每个账户的偏离度。

    Returns:
        {
            "accounts": {
                "安全账户": {
                    "amount": float, "actual": float, "target": float,
                    "absolute_deviation": float, "relative_deviation": float,
                    "status": str, "direction": str
                }, ...
            },
            "total": float,
            "alerts": [str, ...]
        }
    """
    total = portfolio_data["total"]
    if total <= 0:
        return {"accounts": {}, "total": 0, "alerts": ["总资产为0"]}

    result = {"accounts": {}, "total": total, "alerts": []}

    for acc_name, target_pct in TARGET_ALLOCATIONS.items():
        acc_data = portfolio_data["accounts"].get(acc_name, {"amount": 0})
        actual = acc_data["amount"] / total
        abs_dev = abs(actual - target_pct)
        rel_dev = abs_dev / target_pct if target_pct > 0 else 0

        if rel_dev > 0.40:
            status = "紧急"
        elif rel_dev > 0.20:
            status = "关注"
        else:
            status = "正常"

        direction = "超配" if actual > target_pct else "欠配"

        result["accounts"][acc_name] = {
            "amount": acc_data["amount"],
            "actual": round(actual, 4),
            "target": target_pct,
            "absolute_deviation": round(abs_dev, 4),
            "relative_deviation": round(rel_dev, 4),
            "status": status,
            "direction": direction,
            "items_count": len(acc_data.get("items", [])),
        }

        # 生成告警
        if status == "紧急":
            result["alerts"].append(
                f"🔴 {acc_name}{direction} {abs_dev*100:.1f}%（相对偏离 {rel_dev*100:.0f}%），"
                f"实际 {actual*100:.1f}% → 目标 {target_pct*100:.0f}%"
            )
        elif status == "关注":
            result["alerts"].append(
                f"⚠️ {acc_name}{direction} {abs_dev*100:.1f}%（相对偏离 {rel_dev*100:.0f}%），"
                f"实际 {actual*100:.1f}% → 目标 {target_pct*100:.0f}%"
            )

    return result


# ============ 3. 海外指数估值 ============


def _fetch_multpl_series(url: str) -> list:
    """
    从 multpl.com 爬取估值历史数据（月度）。

    Returns:
        [(date_str, float_value), ...] 按时间倒序（最新在前）
    """
    proxy_handler = urllib.request.ProxyHandler({
        "https": PROXY_URL,
        "http": PROXY_URL,
    })
    opener = urllib.request.build_opener(proxy_handler)

    req = urllib.request.Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)",
            "Accept": "text/html",
        },
    )

    try:
        resp = opener.open(req, timeout=20)
        html = resp.read().decode()
    except Exception as e:
        return {"error": str(e), "data": []}

    rows = re.findall(r"<tr[^>]*>(.*?)</tr>", html, re.S)

    data = []
    for r in rows:
        cells = re.findall(r"<td[^>]*>(.*?)</td>", r, re.S)
        if len(cells) >= 2:
            date_str = re.sub(r"<.*?>", "", cells[0]).strip().replace("†", "").strip()
            val_str = (
                re.sub(r"<.*?>|&#x2002;|\n|†|%", "", cells[1])
                .strip()
                .strip()
            )
            try:
                val = float(val_str)
                data.append((date_str, val))
            except ValueError:
                continue

    return {"data": data, "error": None}


def _fetch_multpl_all() -> dict:
    """获取 S&P 500 的 PE/PB/股息率全部数据"""
    result = {}
    for metric, url in MULTPL_URLS.items():
        series = _fetch_multpl_series(url)
        result[metric] = series
    return result


def get_us_index_valuation() -> dict:
    """
    获取标普500估值数据（含历史分位）。

    Returns:
        {
            "标普500": {
                "PE": float, "PE分位_10年": float, "PE分位_5年": float, "PE分位_3年": float,
                "PB": float, "PB分位_10年": float, "PB分位_5年": float,
                "股息率": str, "估值等级_PE": str, "估值等级_PB": str,
                "date": str
            },
            "status": "ok"/"error",
            "error": str or None
        }
    """
    raw = _fetch_multpl_all()

    # 检查错误
    for metric, series in raw.items():
        if series.get("error"):
            return {"status": "error", "error": f"{metric} 获取失败: {series['error']}"}
        if not series.get("data"):
            return {"status": "error", "error": f"{metric} 无数据"}

    result = {"标普500": {}, "status": "ok", "error": None}

    # PE
    pe_data = raw["PE"]["data"]
    if pe_data:
        current_pe = pe_data[0][1]
        pe_values = [v for _, v in pe_data]
        n = len(pe_values)
        result["标普500"]["PE"] = current_pe
        result["标普500"]["PE分位_10年"] = _calc_percentile(pe_values[:120], current_pe) if n >= 120 else _calc_percentile(pe_values, current_pe)
        result["标普500"]["PE分位_5年"] = _calc_percentile(pe_values[:60], current_pe) if n >= 60 else _calc_percentile(pe_values, current_pe)
        result["标普500"]["PE分位_3年"] = _calc_percentile(pe_values[:36], current_pe) if n >= 36 else _calc_percentile(pe_values, current_pe)
        result["标普500"]["估值等级_PE"] = _get_valuation_grade(result["标普500"]["PE分位_10年"])
        result["标普500"]["date"] = _parse_multpl_date(pe_data[0][0])

    # PB
    pb_data = raw["PB"]["data"]
    if pb_data:
        current_pb = pb_data[0][1]
        pb_values = [v for _, v in pb_data]
        n = len(pb_values)
        result["标普500"]["PB"] = current_pb
        result["标普500"]["PB分位_10年"] = _calc_percentile(pb_values[:120], current_pb) if n >= 120 else _calc_percentile(pb_values, current_pb)
        result["标普500"]["PB分位_5年"] = _calc_percentile(pb_values[:60], current_pb) if n >= 60 else _calc_percentile(pb_values, current_pb)
        result["标普500"]["估值等级_PB"] = _get_valuation_grade(result["标普500"]["PB分位_10年"])

    # 股息率
    div_data = raw["股息率"]["data"]
    if div_data:
        result["标普500"]["股息率"] = f"{div_data[0][1]:.2f}%"
        div_values = [v for _, v in div_data]
        result["标普500"]["股息率分位_10年"] = _calc_percentile(div_values[:120], div_data[0][1]) if len(div_values) >= 120 else _calc_percentile(div_values, div_data[0][1])

    return result


def get_hk_index_valuation() -> dict:
    """
    获取港股/中概相关指数估值（通过中证指数官网）。

    Returns:
        {
            "中国互联网": {"PE": float, "滚动PE": float, "股息率": float, "date": str},
            ...
            "status": "ok"/"error"
        }
    """
    # 清除代理（akshare 访问国内网站）
    for k in ["http_proxy", "https_proxy", "HTTP_PROXY", "HTTPS_PROXY"]:
        os.environ.pop(k, None)

    import akshare as ak

    result = {"status": "ok", "error": None}

    for name, code in CSINDEX_HK_MAP.items():
        try:
            df = ak.stock_zh_index_value_csindex(symbol=code)
            if df is not None and len(df) > 0:
                row = df.iloc[-1]
                result[name] = {
                    "代码": code,
                    "PE": float(row.get("市盈率1", 0)) if row.get("市盈率1") else None,
                    "滚动PE": float(row.get("市盈率2", 0)) if row.get("市盈率2") else None,
                    "股息率": float(row.get("股息率1", 0)) if row.get("股息率1") else None,
                    "预期股息率": float(row.get("股息率2", 0)) if row.get("股息率2") else None,
                    "date": str(row.get("日期", "")),
                }
        except Exception as e:
            result[name] = {"error": str(e)}

    return result


# ============ 4. 综合分析 ============


def get_portfolio_analysis() -> dict:
    """
    投资组合综合分析：偏离度 + 海外估值。

    Returns:
        {
            "deviation": {...},      # 偏离度分析
            "us_valuation": {...},   # 标普500估值
            "hk_valuation": {...},   # 港股估值
            "summary": str,          # 文字总结
        }
    """
    # 1. 偏离度
    portfolio = read_portfolio_from_excel()
    deviation = calculate_deviation(portfolio)

    # 2. 美股估值
    us_val = get_us_index_valuation()

    # 3. 港股估值
    hk_val = get_hk_index_valuation()

    # 4. 汇总
    summary_parts = []
    summary_parts.append(f"总资产: ¥{portfolio['total']:,.0f}")
    summary_parts.append("")

    # 偏离度摘要
    for acc_name, acc_data in deviation["accounts"].items():
        actual_pct = acc_data["actual"] * 100
        target_pct = acc_data["target"] * 100
        direction = acc_data["direction"]
        summary_parts.append(
            f"  {acc_name}: {actual_pct:.1f}% → 目标 {target_pct:.0f}% ({direction})"
        )

    # 告警
    if deviation["alerts"]:
        summary_parts.append("")
        for alert in deviation["alerts"]:
            summary_parts.append(alert)

    # 美股估值摘要
    if us_val.get("status") == "ok" and "标普500" in us_val:
        sp = us_val["标普500"]
        summary_parts.append("")
        summary_parts.append(f"标普500: PE={sp.get('PE', 'N/A')} ({sp.get('估值等级_PE', 'N/A')})")
        summary_parts.append(f"  PB={sp.get('PB', 'N/A')}, 股息率={sp.get('股息率', 'N/A')}")

    # 港股估值摘要
    if hk_val.get("status") == "ok":
        summary_parts.append("")
        summary_parts.append("港股/中概:")
        for name in CSINDEX_HK_MAP:
            if name in hk_val and "error" not in hk_val[name]:
                d = hk_val[name]
                summary_parts.append(
                    f"  {name}: PE={d.get('PE', 'N/A')}, 滚动PE={d.get('滚动PE', 'N/A')}"
                )

    return {
        "deviation": deviation,
        "us_valuation": us_val,
        "hk_valuation": hk_val,
        "portfolio": {
            "total": portfolio["total"],
            "accounts": {k: v["amount"] for k, v in portfolio["accounts"].items()},
        },
        "summary": "\n".join(summary_parts),
    }


# ============ 5. 格式化输出 ============


def format_deviation_report(deviation: dict) -> str:
    """格式化偏离度报告"""
    lines = ["📊 四账户偏离度报告", "=" * 40]

    if deviation["alerts"]:
        lines.append("")
        lines.append("⚠️ 告警:")
        for alert in deviation["alerts"]:
            lines.append(f"  {alert}")

    lines.append("")
    lines.append(f"{'账户':<12} {'实际':>6} {'目标':>6} {'偏离':>6} {'相对':>6} {'状态'}")
    lines.append("-" * 55)

    for acc_name, acc in deviation["accounts"].items():
        lines.append(
            f"{acc_name:<10} {acc['actual']*100:>5.1f}% {acc['target']*100:>5.0f}% "
            f"{acc['absolute_deviation']*100:>5.1f}% {acc['relative_deviation']*100:>5.0f}% "
            f"{acc['status']}({acc['direction']})"
        )

    lines.append("")
    lines.append(f"总资产: ¥{deviation['total']:,.0f}")
    return "\n".join(lines)


def format_us_valuation_report(us_val: dict) -> str:
    """格式化美股估值报告"""
    if us_val.get("status") != "ok":
        return f"❌ 美股估值获取失败: {us_val.get('error', '未知错误')}"

    sp = us_val.get("标普500", {})
    lines = ["🇺🇸 标普500估值", "=" * 40]
    lines.append(f"日期: {sp.get('date', 'N/A')}")
    lines.append(f"PE (TTM): {sp.get('PE', 'N/A')}")
    lines.append(f"  10年分位: {sp.get('PE分位_10年', 'N/A')}% → {sp.get('估值等级_PE', 'N/A')}")
    lines.append(f"  5年分位: {sp.get('PE分位_5年', 'N/A')}%")
    lines.append(f"  3年分位: {sp.get('PE分位_3年', 'N/A')}%")
    lines.append(f"PB: {sp.get('PB', 'N/A')}")
    lines.append(f"  10年分位: {sp.get('PB分位_10年', 'N/A')}% → {sp.get('估值等级_PB', 'N/A')}")
    lines.append(f"  5年分位: {sp.get('PB分位_5年', 'N/A')}%")
    lines.append(f"股息率: {sp.get('股息率', 'N/A')}")
    return "\n".join(lines)


def format_hk_valuation_report(hk_val: dict) -> str:
    """格式化港股估值报告"""
    lines = ["🇭🇰 港股/中概估值（中证指数）", "=" * 40]

    for name in CSINDEX_HK_MAP:
        if name in hk_val:
            d = hk_val[name]
            if "error" in d:
                lines.append(f"  {name}: 获取失败 ({d['error']})")
            else:
                lines.append(f"  {name}:")
                lines.append(f"    PE={d.get('PE', 'N/A')}, 滚动PE={d.get('滚动PE', 'N/A')}")
                lines.append(f"    股息率={d.get('股息率', 'N/A')}%, 预期={d.get('预期股息率', 'N/A')}%")
                lines.append(f"    日期: {d.get('date', 'N/A')}")

    return "\n".join(lines)


# ============ CLI ============

if __name__ == "__main__":
    print("=" * 50)
    print("投资组合分析")
    print("=" * 50)

    # 1. 偏离度
    print("\n--- 偏离度分析 ---")
    portfolio = read_portfolio_from_excel()
    deviation = calculate_deviation(portfolio)
    print(format_deviation_report(deviation))

    # 2. 美股估值
    print("\n--- 美股估值 ---")
    us_val = get_us_index_valuation()
    print(format_us_valuation_report(us_val))

    # 3. 港股估值
    print("\n--- 港股/中概估值 ---")
    hk_val = get_hk_index_valuation()
    print(format_hk_valuation_report(hk_val))

    print("\n" + "=" * 50)
